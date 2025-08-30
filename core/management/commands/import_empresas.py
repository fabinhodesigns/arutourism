# core/management/commands/import_empresas.py
import re
from pathlib import Path
from django.core.management.base import BaseCommand, CommandError
from django.contrib.auth.models import User
from django.db import transaction
from openpyxl import load_workbook

from core.models import Categoria, Empresa

# defaults rígidos (imagem e localização obrigatórias)
DEFAULT_LAT = "-28.937100"
DEFAULT_LNG = "-49.484000"
DEFAULT_EMAIL = "sem-email@exemplo.com"
DEFAULT_TEL = "(00)0000-0000"
DEFAULT_IMAGEM_PATH = Path("media/placeholders/sem_imagem.png")

def norm_doc(doc: str) -> str:
    if not doc:
        return ""
    return re.sub(r"\D", "", str(doc))

def norm_str(s):
    if s is None:
        return ""
    return str(s).strip()

def first_non_empty(*vals):
    for v in vals:
        v = norm_str(v)
        if v:
            return v
    return ""

class Command(BaseCommand):
    help = "Importa empresas de um arquivo Excel (XLSX) e cria categorias automaticamente"

    def add_arguments(self, parser):
        parser.add_argument("--file", required=True, help="Caminho do arquivo XLSX (ex.: /caminho/ARARANGUA.xlsx)")
        parser.add_argument("--usuario", default=None, help="username dono dos registros (default: primeiro superuser)")
        parser.add_argument("--cidade", default="Araranguá", help="Cidade padrão quando ausente (default: Araranguá)")

    def get_owner_user(self, username_opt):
        if username_opt:
            u = User.objects.filter(username=username_opt).first()
            if not u:
                raise CommandError(f"Usuário '{username_opt}' não encontrado.")
            return u
        # pega um superuser qualquer
        u = User.objects.filter(is_superuser=True).first()
        if not u:
            # fallback: primeiro user existente
            u = User.objects.first()
        if not u:
            raise CommandError("Nenhum usuário encontrado. Crie um usuário antes do import.")
        return u

    @transaction.atomic
    def handle(self, *args, **options):
        xlsx_path = Path(options["file"]).expanduser().resolve()
        if not xlsx_path.exists():
            raise CommandError(f"Arquivo não encontrado: {xlsx_path}")

        owner = self.get_owner_user(options["usuario"])
        cidade_default = options["cidade"]

        wb = load_workbook(filename=str(xlsx_path))
        # pega a primeira sheet
        ws = wb[wb.sheetnames[0]]

        # mapeia cabeçalhos (case-insensitive, ignora acentos/barras)
        header_row = None
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            header_row = [norm_str(c).lower() for c in row]
            header_map = {h: idx for idx, h in enumerate(header_row) if h}
            start_data_row = i + 1
            break

        def col(*aliases):
            for a in aliases:
                for h, idx in header_map.items():
                    if a in h:
                        return idx
            return None

        idx_cnpj       = col("cnpj")
        idx_categoria  = col("ramo", "categoria", "ramo atividade")
        idx_nome       = col("nome")
        idx_endereco   = col("endereço", "endereco")
        idx_telefone   = col("telefone")
        idx_contato    = col("contato direto", "contato")
        idx_digital    = col("digital", "site", "website")
        idx_cadastrur  = col("cadas", "cadastur")
        idx_maps       = col("maps", "google")
        idx_app        = col("app")
        # latitude/longitude se existir na planilha
        idx_lat        = col("lat")
        idx_lng        = col("lng", "long", "longitude")
        # email se existir
        idx_email      = col("email")

        criadas, atualizadas, puladas = 0, 0, 0

        # garante placeholder de imagem
        if not DEFAULT_IMAGEM_PATH.exists():
            raise CommandError(f"Imagem placeholder não encontrada em {DEFAULT_IMAGEM_PATH}. Crie o arquivo antes de rodar.")

        for row in ws.iter_rows(min_row=start_data_row, values_only=True):
            if row is None:
                continue

            nome = first_non_empty(row[idx_nome] if idx_nome is not None else "")
            if not nome:
                puladas += 1
                continue

            cat_raw = first_non_empty(row[idx_categoria] if idx_categoria is not None else "")
            cat_nome = cat_raw.strip() if cat_raw else "Sem categoria"
            categoria, _ = Categoria.objects.get_or_create(nome=cat_nome)

            cnpj = norm_doc(row[idx_cnpj]) if idx_cnpj is not None else ""
            cadastrur = norm_str(row[idx_cadastrur]) if idx_cadastrur is not None else ""

            endereco_full = norm_str(row[idx_endereco]) if idx_endereco is not None else ""
            telefone = norm_str(row[idx_telefone]) if idx_telefone is not None else ""
            contato_direto = norm_str(row[idx_contato]) if idx_contato is not None else ""
            digital = norm_str(row[idx_digital]) if idx_digital is not None else ""
            maps_url = norm_str(row[idx_maps]) if idx_maps is not None else ""
            app_url = norm_str(row[idx_app]) if idx_app is not None else ""
            email = norm_str(row[idx_email]) if idx_email is not None else ""

            # localização
            lat = norm_str(row[idx_lat]) if idx_lat is not None else ""
            lng = norm_str(row[idx_lng]) if idx_lng is not None else ""
            if not lat: lat = DEFAULT_LAT
            if not lng: lng = DEFAULT_LNG

            # normaliza telefone
            tel_digits = re.sub(r"\D", "", telefone) if telefone else ""
            sem_telefone = False
            if not tel_digits:
                tel_digits = re.sub(r"\D", "", DEFAULT_TEL)
                sem_telefone = True

            sem_email = False
            if not email:
                email = DEFAULT_EMAIL
                sem_email = True

            # Busca por (user, nome) para atualizar ou criar
            empresa = Empresa.objects.filter(user=owner, nome=nome).first()
            created = False
            if not empresa:
                empresa = Empresa(user=owner, nome=nome)
                created = True

            # preenche campos
            empresa.categoria = categoria
            empresa.cnpj = cnpj or None
            empresa.cadastrur = cadastrur or None

            empresa.endereco_full = endereco_full
            # mantém granular em branco (pode-se parsear depois)
            empresa.rua = empresa.rua or ""
            empresa.numero = empresa.numero or ""
            empresa.bairro = empresa.bairro or ""
            empresa.cidade = empresa.cidade or cidade_default
            empresa.cep = empresa.cep or ""

            empresa.telefone = tel_digits
            empresa.sem_telefone = sem_telefone
            empresa.email = email
            empresa.sem_email = sem_email
            empresa.contato_direto = contato_direto or None

            # presença digital
            empresa.site = digital or empresa.site
            empresa.digital = digital or None
            empresa.maps_url = maps_url or None
            empresa.app_url = app_url or None

            # localização obrigatória (defaults já garantem)
            empresa.latitude = lat
            empresa.longitude = lng

            # imagem obrigatória: se vazio no BD, usa placeholder
            if not getattr(empresa, "imagem", None):
                empresa.imagem.name = str(DEFAULT_IMAGEM_PATH).replace("\\", "/")  # usa o arquivo local como default

            empresa.save()
            if created:
                criadas += 1
            else:
                atualizadas += 1

        self.stdout.write(self.style.SUCCESS(
            f"Import finalizado: criadas={criadas}, atualizadas={atualizadas}, puladas={puladas}"
        ))