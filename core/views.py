from __future__ import annotations
import unicodedata

import csv
import io
import re
from io import BytesIO, StringIO
from urllib.parse import urlparse, parse_qs
from .forms import ProfileForm, CpfUpdateForm, StartResetByCpfForm, CustomLoginForm, EmpresaForm, UserRegistrationForm, TagForm
from .models import PerfilUsuario, Empresa, ImagemEmpresa
from django.contrib.auth import authenticate
from .models import Tag
from django.contrib.admin.views.decorators import staff_member_required
import json

from .forms import AvaliacaoForm
from .models import Avaliacao
from django.db import IntegrityError

import logging
logger = logging.getLogger(__name__)

from django.views.decorators.http import require_http_methods
from django.contrib import messages
from django.contrib.auth import login as auth_login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.models import User
from django.core.paginator import Paginator
from django.db import transaction, IntegrityError
from django.db.models import Q, Avg, Count, Prefetch
from django.http import Http404, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import render_to_string
from django.utils.decorators import method_decorator
from django.views.decorators.http import require_GET, require_POST
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import csv, io

from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.forms import PasswordChangeForm, PasswordResetForm
from django.urls import reverse, reverse_lazy
from django.core.mail import send_mail  
from django.utils import timezone

from .models import PerfilUsuario, Empresa, Tag

from core.models import Tag, Empresa

# ============================================================

# nomes aceitos (mantém os seus sinônimos)
COLUMN_ALIASES = {
    # --- Modelo Padrão (Mantido para compatibilidade) ---
    "nome": ["nome", "razão social", "razao social"],
    "categoria": ["ramo atividade", "ramo de atividade", "categoria", "ramo"],
    "bairro": ["bairro"], "rua": ["endereço", "endereco", "logouro", "endereço completo"],
    "numero": ["número", "numero", "nº"], "cidade": ["cidade", "municipio", "município"],
    "cep": ["cep", "c.e.p."], "telefone": ["telefone", "fone", "whatsapp"],
    "contato": ["contato direto", "contato"], "descricao": ["descrição", "descricao", "observacao", "obs"],
    # ... outros do modelo padrão

    # --- MAPEAMENTO COMPLETO PARA GOOGLE CONTACTS ---
    # Nomes
    "google_nome_contato": ["Name"],
    "google_tag": ["First Name"], 
    "google_nome_do_meio": ["Middle Name"],
    "google_sobrenome": ["Last Name"],
    "google_apelido": ["Nickname"],
    "google_file_as": ["File As"],
    
    # Organização
    "google_nome_empresa": ["Organization Name"],
    "google_cargo": ["Organization Title"],
    "google_departamento": ["Organization Department"],
    
    # Contatos
    "google_email_1": ["E-mail 1 - Value"],
    "google_email_2": ["E-mail 2 - Value"],
    "google_telefone_1": ["Phone 1 - Value"],
    "google_telefone_2": ["Phone 2 - Value"],
    "google_telefone_3": ["Phone 3 - Value"],
    
    # Endereços
    "google_endereco_formatado": ["Address 1 - Formatted"],
    "google_rua": ["Address 1 - Street"],
    "google_cidade": ["Address 1 - City"],
    "google_estado": ["Address 1 - Region"],
    "google_cep": ["Address 1 - Postal Code"],
    "google_pais": ["Address 1 - Country"],
    
    # Websites
    "google_website_1": ["Website 1 - Value"],
    "google_website_2": ["Website 2 - Value"],
    
    # Outros
    "google_aniversario": ["Birthday"],
    "google_notas": ["Notes"],
    
    "google_custom_1_label": ["Custom Field 1 - Label"], 
    "google_custom_1_value": ["Custom Field 1 - Value"], 
    "google_custom_2_label": ["Custom Field 2 - Label"],
    "google_custom_2_value": ["Custom Field 2 - Value"],
    "google_custom_3_label": ["Custom Field 3 - Label"],
    "google_custom_3_value": ["Custom Field 3 - Value"],
}

TEMPLATE_HEADERS = [
    ("CNPJ",                          False, "cnpj"),
    ("CATEGORIA (RAMO ATIVIDADE)",    False, "categoria"),
    ("NOME",                          False, "nome"),
    ("BAIRRO",                        False, "bairro"),
    ("ENDEREÇO COMPLETO",             False, "endereco"),
    ("TELEFONE",                      False, "telefone"),
    ("CONTATO DIRETO",                False, "contato"),
    ("DIGITAL (site/redes)",          False, "digital"),
    ("CADASTUR",                      False, "cadastrur"),
    ("MAPS (link)",                   False, "maps"),
    ("APP",                           False, "app"),
    ("DESCRIÇÃO",                     False, "descricao"),
    ("NÚMERO",                        False, "numero"),
    ("CEP",                           False, "cep"),
    ("CIDADE",                        False, "cidade"),
]

LOREM_DEFAULT = (
    "Descrição ainda não informada. Este estabelecimento está em processo de "
    "complementação de dados. Se você é o responsável, atualize as informações."
)

DEFAULT_DESC = (
    "Descrição ainda não informada. Este estabelecimento está em processo de "
    "complementação de dados. Se você é o responsável, atualize as informações."
)

EXPECTED_CANONS = [c for (_label, _req, c) in TEMPLATE_HEADERS]
HUMAN_LABEL_BY_CANON = {c: label for (label, _req, c) in TEMPLATE_HEADERS}

# ============================================================
# Helpers
# ============================================================

def get_base_empresas_queryset():
    return Empresa.objects.select_related('user__perfil').prefetch_related(
        'tags',
        Prefetch('imagens', queryset=ImagemEmpresa.objects.filter(principal=True), to_attr='imagem_principal_list')
    ).annotate(
        avg_nota=Avg('avaliacoes__nota'),
        count_avaliacoes=Count('avaliacoes')
    )
    
def _clip(model_cls, field_name, value):
    """Trunca strings para caber no max_length do campo (se houver)."""
    if value is None:
        return value
    try:
        f = model_cls._meta.get_field(field_name)
    except Exception:
        return value
    if hasattr(f, "max_length") and f.max_length and isinstance(value, str):
        return value[:f.max_length]
    return value


def _maxlen(model_cls, field_name):
    try:
        f = model_cls._meta.get_field(field_name)
        return getattr(f, "max_length", None)
    except Exception:
        return None

def _norm_text(s: str | None) -> str:
    return re.sub(r"[\W_]+", " ", (s or "")).strip().lower()

def _digits(s: str | None) -> str:
    return re.sub(r"\D", "", s or "")

def _extract_latlng_from_maps(url):
    try:
        if not url: return None, None
        u = urlparse(url)
        qs = parse_qs(u.query)
        if "q" in qs and "," in qs["q"][0]:
            lat, lng = qs["q"][0].split(",")[:2]
            return float(lat), float(lng)
        m = re.search(r"@(-?\d+\.\d+),(-?\d+\.\d+)", url)
        if m: return float(m.group(1)), float(m.group(2))
    except Exception:
        pass
    return None, None

def _alias_match(key_norm: str, alias: str) -> bool:
    """Casa de forma flexível: tokens do alias presentes no cabeçalho."""
    a = _norm_text(alias).split()
    k = key_norm.split()
    return all(t in k for t in a) or key_norm == _norm_text(alias)

def _build_header_map(raw_headers):
    mapping = {}
    for idx, raw in enumerate(raw_headers or []):
        key = _norm_text(raw)
        for canon, alts in COLUMN_ALIASES.items():
            if key in (_norm_text(a) for a in alts):
                mapping[idx] = canon
                break
    return mapping

def _read_rows_from_upload(file_obj, filename):
    name = (filename or "").lower()
    if name.endswith(".csv"):
        data = file_obj.read().decode("utf-8", errors="ignore")
        reader = csv.reader(StringIO(data))
        rows = list(reader)
    else:
        from openpyxl import load_workbook
        wb = load_workbook(file_obj, data_only=True)
        ws = wb.active
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append([(c if c is not None else "") for c in row])
    if not rows: 
        return [], []
    headers = [str(c).strip() for c in rows[0]]
    body    = [[str(c).strip() for c in r] for r in rows[1:]]
    return headers, body

def _norm_text(s): 
    return re.sub(r"[\W_]+"," ",(s or "")).strip().lower()

    # XLSX
    from openpyxl import load_workbook
    wb = load_workbook(file_obj, data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([(c if c is not None else "") for c in row])
    if not rows:
        return [], []
    headers = [str(c).strip() for c in rows[0]]
    body = [[str(c).strip() for c in r] for r in rows[1:]]
    return headers, body

def _digits(s: str | None) -> str:
    return re.sub(r"\D", "", s or "")

def _squeeze(s: str | None) -> str:
    """trim + colapsa espaços"""
    return re.sub(r"\s+", " ", (s or "").strip())

def _norm_label(s: str | None) -> str:
    """normaliza rótulo de coluna: lowercase, remove '(...)', colapsa espaços"""
    s = (s or "")
    s = re.sub(r"\([^)]*\)", "", s)
    s = s.replace(":", " ").replace("/", " ")
    s = _squeeze(s).lower()
    return s

def _norm_space_lower(s: str | None) -> str:
    return re.sub(r"[\W_]+", " ", (s or "")).strip().lower()

def _to_float(val, default=None):
    try:
        if val is None or str(val).strip() == "":
            return default
        return float(str(val).replace(",", "."))
    except Exception:
        return default

def _match_header_map(headers: list[str]) -> dict[int, str]:
    result = {}
    for idx, raw in enumerate(headers):
        k = _norm_space_lower(raw)
        for canon, alts in COLUMN_ALIASES.items():
            if k in [_norm_space_lower(a) for a in alts]:
                result[idx] = canon
                break
    return result

def _get_reader(file_obj, filename: str):
    ext = (filename or "").lower()
    if ext.endswith(".csv"):
        data = file_obj.read().decode("utf-8", errors="ignore")
        reader = csv.reader(StringIO(data))
        rows = [[(c or "").strip() for c in r] for r in reader]
        return rows

    from openpyxl import load_workbook
    wb = load_workbook(file_obj, data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([("" if c is None else str(c)).strip() for c in row])
    return rows


def _only_digits(s: str) -> str:
    import re
    return re.sub(r"\D", "", s or "")

def _to_float(val, default=None):
    try:
        if val is None or str(val).strip() == "":
            return default
        return float(str(val).replace(",", "."))
    except Exception:
        return default

def _strip_parens(s: str | None) -> str:
    """Remove QUALQUER sufixo entre parênteses no cabeçalho."""
    if not s:
        return ""
    return re.sub(r"\s*\(.*?\)\s*$", "", str(s)).strip()

def _has_field(model_class, field_name: str) -> bool:
    return any(getattr(f, "name", None) == field_name for f in model_class._meta.get_fields())

def _looks_url(s):
    s = (s or "").strip()
    return s.startswith("http://") or s.startswith("https://")

def _first_url_in_text(s):
    s = (s or "")
    m = re.search(r'(https?://\S+)', s)
    return m.group(1) if m else None

def _wants_json(request):
    h = (request.headers.get("Accept") or "").lower()
    return (
        request.GET.get('ajax') == '1' or  
        request.headers.get("x-requested-with") == "XMLHttpRequest" or
        "application/json" in h
    )

def _ident_kind(ident: str) -> str:
    ident = (ident or "").strip()
    if "@" in ident and "." in ident:
        return "email"
    digits = re.sub(r"\D+", "", ident)
    if len(digits) >= 11:
        return "cpf"
    return "username"

def _parse_row_google(data):
    """
    Extrai e normaliza dados de uma linha do Google Contacts, capturando todos os campos
    não mapeados e adicionando-os à descrição.
    """
    parsed = {}
    extra_descricao_parts = []
    
    # Lista de chaves que já têm um tratamento especial e não devem ser repetidas na descrição.
    handled_keys = {
        'google_nome_empresa', 'google_file_as', 'google_nome_contato', 'google_tag',
        'google_email_1', 'google_email_2', 'google_telefones', 'google_telefone_1', 
        'google_telefone_2', 'google_telefone_3', 'google_website_1', 'google_website_2',
        'google_endereco_formatado', 'google_rua', 'google_cidade', 'google_cep', 'google_notas',
        'google_custom_1_label', 'google_custom_1_value', 'google_custom_2_label', 
        'google_custom_2_value', 'google_custom_3_label', 'google_custom_3_value',
    }

    # --- 1. MAPEAMENTO DE CAMPOS PRINCIPAIS ---
    
    # Nome da Empresa (com fallbacks)
    parsed['nome'] = data.get('google_nome_empresa') or data.get('google_file_as') or data.get('google_nome_contato')
    
    # Tag (Primeiro Nome)
    if data.get('google_tag'):
        parsed['tags'] = [data.get('google_tag')]

    # E-mail (pega o primeiro encontrado)
    parsed['email'] = data.get('google_email_1') or data.get('google_email_2')

    # Telefones (pega o primeiro para o campo principal, os outros vão para a descrição)
    all_phones = [data.get('google_telefone_1', ''), data.get('google_telefone_2', ''), data.get('google_telefone_3', '')]
    valid_phones = [phone for phone in all_phones if phone]
    if valid_phones:
        parsed['telefone'] = _digits(valid_phones[0])
        if len(valid_phones) > 1:
            extra_descricao_parts.append(f"Telefones Adicionais: {', '.join(valid_phones[1:])}")

    # Endereço
    parsed['rua'] = data.get('google_rua')
    parsed['cidade'] = data.get('google_cidade')
    parsed['cep'] = _digits(data.get('google_cep'))

    # Websites (lógica para Instagram/Facebook)
    website_url = data.get('google_website_1', '')
    if 'instagram.com' in website_url: parsed['instagram'] = website_url
    elif 'facebook.com' in website_url: parsed['facebook'] = website_url
    elif website_url: parsed['site'] = website_url

    # Campos Customizados (CNPJ e CADASTUR)
    for i in range(1, 4):
        label = data.get(f'google_custom_{i}_label', '').lower().strip()
        value = data.get(f'google_custom_{i}_value', '').strip()
        if label in ['cpf', 'cnpj', 'cpf ou cnpj']:
            parsed['cnpj'] = _digits(value)
        elif label == 'cadastur':
            parsed['cadastrur'] = "Sim" if value.lower() in ['sim', 's'] else "Não"

    # --- 2. CAPTURA DE TODOS OS CAMPOS NÃO MAPEADOS ---
    
    unmapped_data = []
    for key, value in data.items():
        if key not in handled_keys and value and 'label' not in key:
            # Transforma a chave (ex: 'google_sobrenome') em um rótulo legível (ex: 'Sobrenome')
            label_legivel = key.replace('google_', '').replace('_', ' ').title()
            unmapped_data.append(f"{label_legivel}: {value}")
    
    # --- 3. MONTAGEM DA DESCRIÇÃO FINAL ---
    
    # Começa com o campo "Notas" do Google
    descricao_final = data.get('google_notas', '')
    
    # Adiciona o endereço formatado, se existir
    if data.get('google_endereco_formatado'):
        descricao_final += f"\n\nEndereço Completo: {data.get('google_endereco_formatado')}"

    # Adiciona a lista de telefones extras (se houver)
    if any("Telefones Adicionais" in part for part in extra_descricao_parts):
        descricao_final += "\n" + "\n".join(part for part in extra_descricao_parts if "Telefones Adicionais" in part)
        
    # Adiciona todos os outros campos não mapeados
    if unmapped_data:
        descricao_final += "\n\n--- Outras Informações ---\n"
        descricao_final += "\n".join(unmapped_data)
        
    parsed['descricao'] = descricao_final or DEFAULT_DESC
    
    return parsed

# ============================================================
# Páginas básicas / Auth
# ============================================================

def home(request):
    empresas_list = get_base_empresas_queryset().order_by('-data_cadastro')[:8]
    total_empresas = Empresa.objects.count()

    favorito_ids = []
    if request.user.is_authenticated:
        favorito_ids = request.user.perfil.favoritos.values_list('id', flat=True)

    return render(request, 'home.html', {
        'page_obj': empresas_list,
        'total_empresas': total_empresas,
        'favorito_ids': list(favorito_ids)
    })


def sobre(request):
    return render(request, 'core/sobre.html')

def termo_de_servico(request):
    return render(request, 'core/termo_de_servico.html')

def politica_de_privacidade(request):
    return render(request, 'core/politica_de_privacidade.html')

def register(request):
    if request.user.is_authenticated:
        return redirect('home')

    if request.method == 'POST':
        form = UserRegistrationForm(request.POST)
        if form.is_valid():
            form.save()
            if _wants_json(request):
                return JsonResponse({"ok": True, "redirect": reverse("login")})
            return redirect('login')

        if _wants_json(request):
            return JsonResponse({"ok": False, "errors": form.errors}, status=400)

        for field, errors in form.errors.items():
            for error in errors:
                messages.error(request, f"{field}: {error}")
    else:
        form = UserRegistrationForm()

    return render(request, 'core/register.html', {'form': form})

def login_view(request):
    if request.user.is_authenticated:
        return redirect('home')

    if request.method == 'POST':
        ident = (request.POST.get('identificador') or "").strip()
        pwd = request.POST.get('password') or ""
        
        user_obj = None
        
        if '@' in ident:
            user_obj = User.objects.filter(email__iexact=ident).first()
        elif re.match(r'^\d{11}$|^\d{14}$', re.sub(r'\D', '', ident)):
            digits = re.sub(r'\D', '', ident)
            perfil = PerfilUsuario.objects.select_related("user").filter(cpf_cnpj=digits).first()
            if perfil:
                user_obj = perfil.user
        else:
            user_obj = User.objects.filter(username__iexact=ident).first()

        if user_obj:
            user = authenticate(request, username=user_obj.username, password=pwd)
            
            if user is not None:
                if user.is_active:
                    print(f"[LOGIN] Autenticação manual OK para user.id={user.id}")
                    auth_login(request, user)
                    
                    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                        return JsonResponse({"ok": True, "redirect": reverse("home")})
                    return redirect('home')
                else:
                    print(f"[LOGIN] Falha: usuário '{user.username}' está inativo.")
                    messages.error(request, 'Esta conta está desativada.')
            else:
                print(f"[LOGIN] Falha: Senha incorreta para o usuário '{user_obj.username}'.")
                messages.error(request, 'Por favor, verifique seu identificador e senha.')
        else:
            print(f"[LOGIN] Falha: Nenhum usuário encontrado para o identificador '{ident}'.")
            messages.error(request, 'Por favor, verifique seu identificador e senha.')
        
        form = CustomLoginForm(request.POST)
        
        if request.headers.get('X-Requested-with') == 'XMLHttpRequest':
            error_message = str(list(messages.get_messages(request))[-1])
            return JsonResponse({
                "ok": False, 
                "errors": [{"message": error_message}]
            }, status=400)
            
        return render(request, 'core/login.html', {'form': form})

    form = CustomLoginForm()
    return render(request, 'core/login.html', {'form': form})

def logout_view(request):
    logout(request)
    return redirect('login')

# ============================================================
# Usuários (admin simples)
# ============================================================

def listar_usuarios(request):
    if not request.user.is_authenticated:
        return redirect('login')
    usuarios = User.objects.all() if request.user.is_superuser else User.objects.filter(id=request.user.id)
    return render(request, '/listar_usuarios.html', {'usuarios': usuarios})

def editar_usuario(request, id):
    usuario = get_object_or_404(User, id=id)
    if request.user != usuario and not request.user.is_superuser:
        raise Http404("Você não tem permissão para editar este usuário.")
    if request.method == 'POST':
        form = UserRegistrationForm(request.POST, instance=usuario)
        if form.is_valid():
            form.save()
            return redirect('listar_usuarios')
    else:
        form = UserRegistrationForm(instance=usuario)
    return render(request, '/editar_usuario.html', {'form': form})

def excluir_usuario(request, id):
    usuario = get_object_or_404(User, id=id)
    if request.user != usuario and not request.user.is_superuser:
        raise Http404("Você não tem permissão para excluir este usuário.")
    if request.method == 'POST':
        usuario.delete()
        return redirect('listar_usuarios')
    return render(request, '/excluir_usuario.html', {'usuario': usuario})

@login_required(login_url='/login/')
@require_http_methods(["GET", "POST"])
def cadastrar_empresa(request):
    initial = {"latitude": "-28.937100", "longitude": "-49.484000"}

    wants_json = (
        'application/json' in (request.headers.get('Accept') or '')
        or (request.headers.get('x-requested-with') == 'XMLHttpRequest')
    )

    if request.method == 'GET':
        form = EmpresaForm(initial=initial)
        return render(request, 'core/cadastrar_empresa.html', {
            'form': form,
            'is_editing': False
        })

    form = EmpresaForm(request.POST, request.FILES)
    if form.is_valid():
        empresa = form.save(commit=False)
        empresa.user = request.user
        empresa.save()

        imagem_file = request.FILES.get('imagem_inicial')
        if imagem_file:
            ImagemEmpresa.objects.create(
                empresa=empresa,
                imagem=imagem_file,
                principal=True 
            )

        if wants_json:
            action = 'reset' if 'save_and_add' in request.POST else 'redirect'
            return JsonResponse({
                'ok': True,
                'status': 'success',
                'message': f"Empresa '{empresa.nome}' cadastrada com sucesso!",
                'action': action,
                'redirect_url': reverse('suas_empresas') if action == 'redirect' else ''
            })

        messages.success(request, f"Empresa '{empresa.nome}' cadastrada com sucesso!")
        return redirect('suas_empresas')

    if wants_json:
        html = render_to_string('core/partials/form_errors.html', {'form': form}, request=request)
        return JsonResponse(
            {'ok': False, 'error': 'Erros de validação no formulário.', 'html': html},
            status=400
        )

    return render(request, 'core/cadastrar_empresa.html', {
        'form': form,
        'is_editing': False
    }, status=400)

@login_required(login_url='/login/')
@require_http_methods(["GET", "POST"])
def editar_empresa(request, slug):
    empresa = get_object_or_404(Empresa, slug=slug)

    if empresa.user_id != request.user.id and not request.user.is_superuser:
        raise Http404("Você não tem permissão para editar esta empresa.")

    wants_json = (
        'application/json' in (request.headers.get('Accept') or '')
        or request.headers.get('x-requested-with') == 'XMLHttpRequest'
    )

    if request.method == 'GET':
        form = EmpresaForm(instance=empresa)
    else:
        form = EmpresaForm(request.POST, request.FILES, instance=empresa)
        if form.is_valid():
            form.save()

            novas_imagens = request.FILES.getlist('novas_imagens')
            for file in novas_imagens:
                ImagemEmpresa.objects.create(empresa=empresa, imagem=file)

            principal_id = request.POST.get('imagem_principal')
            if principal_id:
                empresa.imagens.update(principal=False)
                ImagemEmpresa.objects.filter(id=principal_id, empresa=empresa).update(principal=True)

            if wants_json:
                return JsonResponse({'ok': True, 'redirect_url': reverse('suas_empresas')})
            messages.success(request, f"A empresa “{empresa.nome}” foi atualizada com sucesso!")
            return redirect('suas_empresas')

        if wants_json:
            html = render_to_string('core/partials/form_errors.html', {'form': form}, request=request)
            return JsonResponse({'ok': False, 'html': html}, status=400)

    return render(request, 'core/editar_empresa.html', {
        'form': form,
        'empresa': empresa,
    }, status=(400 if request.method == 'POST' else 200))

@login_required
@require_POST
def deletar_imagem_empresa(request, imagem_id):
    imagem = get_object_or_404(ImagemEmpresa, id=imagem_id)
    empresa = imagem.empresa

    if empresa.user != request.user and not request.user.is_superuser:
        return JsonResponse({'status': 'error', 'message': 'Permissão negada.'}, status=403)

    if imagem.principal:
        imagem.delete()
        nova_principal = empresa.imagens.order_by('data_upload').first()
        if nova_principal:
            nova_principal.principal = True
            nova_principal.save()
    else:
        imagem.delete()

    return JsonResponse({'status': 'success', 'message': 'Imagem deletada.'})


@login_required(login_url='/login/')
def suas_empresas(request):
    empresas_list = get_base_empresas_queryset().filter(user=request.user).order_by('-data_cadastro')
    paginator = Paginator(empresas_list, 8)
    page_obj = paginator.get_page(request.GET.get('page') or 1)

    if _wants_json(request):
        html = render_to_string('core/partials/empresas_cards.html', {'page_obj': page_obj}, request=request)
        return JsonResponse({'html': html, 'has_next': page_obj.has_next(),
                             'next_page_number': page_obj.next_page_number() if page_obj.has_next() else None})
    
    return render(request, 'core/suas_empresas.html', {'page_obj': page_obj})


def empresa_detalhe(request, slug):
    empresa = get_object_or_404(
        Empresa.objects.annotate(
            avg_nota=Avg('avaliacoes__nota'),
            count_avaliacoes=Count('avaliacoes__id', distinct=True)
        ).prefetch_related('imagens', 'avaliacoes__user__perfil'), 
        slug=slug
    )

    avaliacao_form = AvaliacaoForm()
    user_ja_avaliou = False
    is_favorito = False

    if request.user.is_authenticated:
        if Avaliacao.objects.filter(empresa=empresa, user=request.user).exists():
            user_ja_avaliou = True

            is_favorito = request.user.perfil.favoritos.filter(slug=slug).exists()

    context = {
        'empresa': empresa,
        'avaliacao_form': avaliacao_form,
        'user_ja_avaliou': user_ja_avaliou,
        'is_favorito': is_favorito, 
    }
    return render(request, 'core/empresa_detalhe.html', context)


def listar_empresas(request):
    empresas = get_base_empresas_queryset()

    q = (request.GET.get('q') or '').strip()
    tag_ids = request.GET.getlist('tag')
    cidade = (request.GET.get('cidade') or '').strip()
    
    if q:
        empresas = empresas.filter(
            Q(nome__icontains=q) |
            Q(descricao__icontains=q) |
            Q(bairro__icontains=q) |
            Q(cidade__icontains=q) |
            Q(tags__nome__icontains=q)
        ).distinct()

    tag_labels = []
    if tag_ids:
        empresas = empresas.filter(tags__id__in=tag_ids).distinct()
        
        selected_tags = Tag.objects.filter(id__in=tag_ids)
        tag_labels = [tag.nome for tag in selected_tags]
    
    if cidade:
        empresas = empresas.filter(cidade__iexact=cidade)

    empresas = empresas.order_by('-id')

    paginator = Paginator(empresas, 12)
    page_obj = paginator.get_page(request.GET.get('page') or 1)

    if _wants_json(request):
        html = render_to_string('core/partials/empresas_cards.html', {'page_obj': page_obj}, request=request)
        return JsonResponse({'html': html, 'has_next': page_obj.has_next(), 'next_page_number': page_obj.next_page_number() if page_obj.has_next() else None})

    filtros_aplicados = { 'q': q, 'tag': tag_ids, 'cidade': cidade }
    filtros_legiveis = { 'q': q or None, 'tag': ", ".join(tag_labels), 'cidade': cidade or None }

    favorito_ids = []
    if request.user.is_authenticated:
        perfil, created = PerfilUsuario.objects.get_or_create(user=request.user)
        favorito_ids = perfil.favoritos.values_list('id', flat=True)

    return render(request, 'core/listar_empresas.html', {
        'page_obj': page_obj,
        'filtros_aplicados': filtros_aplicados,
        'filtros_legiveis': filtros_legiveis,
        'favorito_ids': list(favorito_ids),
    })

def buscar_empresas(request):
    """Rota legada: redireciona para a listagem com os mesmos GETs."""
    return listar_empresas(request)

@login_required
@require_POST
def deletar_avaliacao(request, avaliacao_id):
    avaliacao = get_object_or_404(Avaliacao, id=avaliacao_id)

    if (request.user != avaliacao.user and 
        request.user != avaliacao.empresa.user and 
        not request.user.is_superuser):
        return JsonResponse({'status': 'error', 'message': 'Permissão negada.'}, status=403)

    avaliacao.delete()

    return JsonResponse({'status': 'success', 'message': 'Avaliação removida com sucesso.'})

@require_POST
@login_required
def deletar_empresa(request, slug):
    # Encontra a empresa pelo slug
    empresa = get_object_or_404(Empresa, slug=slug)

    # Verifica permissão: somente o dono ou um superusuário pode deletar
    if not request.user.is_superuser and empresa.user != request.user:
        messages.error(request, "Você não tem permissão para deletar esta empresa.")
        return redirect('suas_empresas') # Ou outra página de erro

    nome_empresa = empresa.nome
    empresa.delete()

    messages.success(request, f'A empresa "{nome_empresa}" foi deletada com sucesso.')
    # Redireciona para a lista de empresas do usuário após a exclusão
    return redirect('suas_empresas')

@require_GET
def filtros_empresas(request):
    """Retorna tags e cidades em JSON."""
    tags = list(Tag.objects.order_by('nome').values('id', 'nome'))
    
    cidades = list(
        Empresa.objects.exclude(cidade__isnull=True).exclude(cidade__exact='')
        .order_by('cidade').values_list('cidade', flat=True).distinct()
    )
    return JsonResponse({'tags': tags, 'cidades': list(cidades)})

@require_GET
def download_template_empresas(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Empresas"

    headers = [
        "CNPJ",
        "CATEGORIA",
        "NOME",
        "BAIRRO",
        "ENDEREÇO COMPLETO",
        "TELEFONE",
        "CONTATO DIRETO",
        "DIGITAL (site/redes)",
        "CADASTUR",
        "MAPS (link)",
        "APP",
        "DESCRIÇÃO",
        "NÚMERO",
        "CEP",
        "CIDADE",
        "HORÁRIO SEMANA (seg-sex)",
        "HORÁRIO SÁBADO",
        "HORÁRIO DOMINGO",
        "OBSERVAÇÕES HORÁRIO",
    ]
    ws.append(headers)

    exemplo = [
        "12.345.678/0001-99",
        "Pousada",
        "Pousada Azul",
        "Centro",
        "Av. Central, 123",
        "(48) 99999-9999",
        "Maria (WhatsApp)",
        "site: https://exemplo.com / insta: @pousada",
        "123456789/0123-4",
        "https://maps.google.com/?q=-28.93,-49.48",
        "",
        "Hotel aconchegante com café da manhã.",
        "123",
        "88900-000",
        "Araranguá",
        "09:00 - 18:00",
        "09:00 - 12:00",
        "Fechado",
        "Fechado para almoço das 12h às 13h",
    ]
    ws.append(exemplo)

    for i in range(1, len(headers)+1):
        ws.column_dimensions[get_column_letter(i)].width = 28

    bio = BytesIO()
    wb.save(bio); bio.seek(0)
    return HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="modelo_empresas.xlsx"'},
    )

@login_required
@require_POST
@transaction.atomic
def importar_empresas_arquivo(request):
    # Contadores para o relatório final
    criados = 0
    atualizados = 0
    sem_alteracao = 0
    erros = 0
    msgs = []

    # --- LOG INICIAL ---
    print("\n--- INICIANDO NOVA IMPORTAÇÃO DE ARQUIVO ---")

    try:
        up = request.FILES.get("arquivo")
        if not up:
            print("[ERRO] Nenhum arquivo foi enviado.")
            return JsonResponse({"ok": False, "error": "Envie um arquivo .xlsx ou .csv."}, status=400)

        print(f"Arquivo recebido: {up.name}")
        headers_raw, rows = _read_rows_from_upload(up, up.name)
        if not rows:
            print("[ERRO] Arquivo está vazio ou não pôde ser lido.")
            return JsonResponse({"ok": False, "error": "Arquivo vazio ou em formato inválido."}, status=400)

        # DETECÇÃO DO FORMATO
        is_google_format = len(headers_raw) > 30
        print(f"Detectado formato: {'Google Contacts' if is_google_format else 'Padrão'}")
        
        header_map = _build_header_map(headers_raw)
        print(f"Mapeamento de cabeçalho: {header_map}")
        
        print(f"\nIniciando processamento de {len(rows)} linhas...")
        for line_no, r in enumerate(rows, start=2):
            print(f"\n--- Processando Linha {line_no} ---")
            data = {}
            for idx, canon in header_map.items():
                if idx < len(r): data[canon] = (r[idx] or "").strip()

            if is_google_format:
                dados_empresa = _parse_row_google(data)
            else:
                dados_empresa = _parse_row_padrao(data)
            
            print(f"Dados extraídos da linha: {dados_empresa}")

            nome = dados_empresa.get('nome')
            if not nome:
                erros += 1
                msg = f"Linha {line_no}: O campo 'nome' é obrigatório e não foi encontrado."
                msgs.append(msg)
                print(f"[ERRO] {msg}")
                continue

            # Extrai chaves de busca
            telefone = dados_empresa.get('telefone', '')
            cnpj = dados_empresa.get('cnpj', '')
            tag_names = dados_empresa.pop('tags', [])

            # Lógica de verificação por prioridade
            empresa_existente = None
            if telefone: empresa_existente = Empresa.objects.filter(telefone=telefone).first()
            if not empresa_existente and cnpj: empresa_existente = Empresa.objects.filter(cnpj=cnpj).first()
            if not empresa_existente and nome: empresa_existente = Empresa.objects.filter(nome__iexact=nome).first()

            if empresa_existente:
                print(f"Empresa encontrada no banco: '{empresa_existente.nome}' (ID: {empresa_existente.id}). Verificando atualizações...")
                emp = empresa_existente
                alterado = False
                for campo, valor in dados_empresa.items():
                    if getattr(emp, campo) != valor:
                        print(f"  -> Campo '{campo}' alterado de '{getattr(emp, campo)}' para '{valor}'")
                        setattr(emp, campo, valor)
                        alterado = True
                
                if alterado:
                    emp.save()
                    atualizados += 1
                    print("  -> Empresa ATUALIZADA.")
                else:
                    sem_alteracao += 1
                    print("  -> Nenhuma alteração encontrada.")
            else:
                print(f"Empresa '{nome}' não encontrada. Tentando criar novo registro...")
                try:
                    # Adiciona dados obrigatórios que não podem ser nulos
                    dados_empresa['user'] = request.user
                    # Exemplo: Se 'descricao' for obrigatório e não vier, usa o DEFAULT_DESC
                    dados_empresa.setdefault('descricao', DEFAULT_DESC)
                    
                    emp = Empresa.objects.create(**dados_empresa)
                    criados += 1
                    print(f"  -> Empresa '{nome}' CRIADA com sucesso (ID: {emp.id}).")
                    
                    if tag_names:
                        tag_objects = [Tag.objects.get_or_create(nome=name.strip())[0] for name in tag_names]
                        emp.tags.set(tag_objects)
                        print(f"  -> Tags associadas: {[tag.nome for tag in tag_objects]}")

                except Exception as e:
                    erros += 1
                    msg = f"Linha {line_no} ('{nome}'): Erro ao criar no banco: {e}"
                    msgs.append(msg)
                    print(f"[ERRO GRAVE] {msg}")
                    continue

        print("\n--- FIM DA IMPORTAÇÃO ---")
        print(f"Resultado: {criados} criados, {atualizados} atualizados, {sem_alteracao} sem alteração, {erros} erros.")
        
        return JsonResponse({ "ok": True, "criados": criados, "atualizados": atualizados, "sem_alteracao": sem_alteracao, "erros": erros, "mensagens": msgs })
        
    except Exception as e:
        # Este 'except' pega erros maiores que acontecem ANTES do loop (ex: erro ao ler o arquivo)
        print(f"[ERRO CATASTRÓFICO] Erro inesperado durante a importação: {e}")
        logger.error(f"Erro catastrófico na importação: {e}", exc_info=True)
        return JsonResponse({"ok": False, "error": f"Erro inesperado: {e}"}, status=500)
    

@login_required
def perfil(request):
    perfil, _ = PerfilUsuario.objects.get_or_create(
        user=request.user,
        defaults={"cpf_cnpj": "", "full_name": request.user.get_full_name() or request.user.username},
    )

    empresas_qtd = Empresa.objects.filter(user=request.user).count()
    entrou_fmt = timezone.localtime(request.user.date_joined).strftime("%m/%Y")

    if request.method == 'POST':
        form = ProfileForm(request.POST, request.FILES, instance=perfil, user=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, "Perfil atualizado com sucesso!")
            return redirect('perfil')
    else:
        form = ProfileForm(instance=perfil, user=request.user)

    cpf_form = CpfUpdateForm(request.user) 

    return render(request, 'core/perfil.html', {
        'form': form,
        'cpf_form': cpf_form,
        'empresas_qtd': empresas_qtd,
        'entrou_fmt': entrou_fmt,
    })

@login_required
def trocar_senha(request):
    if request.method == 'POST':
        form = PasswordChangeForm(user=request.user, data=request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user) 
            messages.success(request, "Senha atualizada com sucesso!")
            return redirect('perfil')
    else:
        form = PasswordChangeForm(user=request.user)

    return render(request, 'core/password_change.html', {'form': form})


def esqueci_senha_email(request):
    if request.method == 'POST':
        form = PasswordResetForm(request.POST)
        if form.is_valid():
            form.save(
                request=request,
                use_https=request.is_secure(),
                email_template_name='core/email_password_reset.txt',
                html_email_template_name='core/email_password_reset.html',
                subject_template_name='core/email_password_reset_subject.txt',
                from_email=None
            )
            messages.success(request, "Se o e-mail existir no sistema, você receberá um link para alterar sua senha.")
            return redirect('perfil')
    else:
        form = PasswordResetForm()
    return render(request, 'core/password_reset_email.html', {'form': form})


def esqueci_senha_cpf(request):
    msg_privacidade = (
        "Por segurança, se o CPF/CNPJ existir e estiver vinculado a um e-mail, "
        "enviaremos um link de redefinição para esse e-mail."
    )
    if request.method == 'POST':
        form = StartResetByCpfForm(request.POST)
        if form.is_valid():
            cpf_digits = form.cleaned_data['cpf_cnpj']  
            perfil = (PerfilUsuario.objects
                      .select_related('user')
                      .filter(cpf_cnpj=cpf_digits)
                      .first())

            if perfil and perfil.user.email:
                prf = PasswordResetForm({'email': perfil.user.email})
                if prf.is_valid():
                    prf.save(
                        request=request,
                        use_https=request.is_secure(),
                        email_template_name='core/email_password_reset.txt',
                        html_email_template_name='core/email_password_reset.html',
                        subject_template_name='core/email_password_reset_subject.txt',
                        from_email=None
                    )
                    logger.info("Password reset enviado por CPF %s → %s", cpf_digits, perfil.user.email)
                else:
                    logger.warning("PasswordResetForm inválido para e-mail %s (CPF %s)", perfil.user.email, cpf_digits)
            else:
                logger.info("CPF/CNPJ não localizado ou sem e-mail cadastrado: %s", cpf_digits)

            messages.success(request, msg_privacidade)
            return redirect('perfil')
    else:
        form = StartResetByCpfForm()

    return render(request, 'core/password_reset_cpf.html', {'form': form})

@login_required
@require_POST
def perfil_alterar_cpf(request):
    with transaction.atomic():
        perfil = (PerfilUsuario.objects
                  .select_for_update()
                  .select_related('user')
                  .get(user=request.user))

        form = CpfUpdateForm(request.user, request.POST)
        if not form.is_valid():
            form_profile = ProfileForm(instance=perfil, user=request.user)
            empresas_qtd = Empresa.objects.filter(user=request.user).count()
            entrou_fmt = timezone.localtime(request.user.date_joined).strftime("%m/%Y")
            return render(request, 'core/perfil.html', {
                'form': form_profile,
                'cpf_form': form,
                'empresas_qtd': empresas_qtd,
                'entrou_fmt': entrou_fmt,
            }, status=400)

        novo_cpf = form.cleaned_data["cpf_cnpj"]
        if _only_digits(perfil.cpf_cnpj) == novo_cpf:
            messages.info(request, "O CPF informado é igual ao atual.")
            return redirect('perfil')

        try:
            perfil.cpf_cnpj = novo_cpf 
            perfil.save(update_fields=["cpf_cnpj"])
        except IntegrityError:
            messages.error(request, "Este CPF já está em uso por outra conta.")
            return redirect('perfil')

    messages.success(request, "CPF atualizado com sucesso!")
    return redirect('perfil')


def page_not_found(request, exception):
    return render(request, 'core/404.html', status=500)

def server_error(request):
    return render(request, 'core/500.html', status=500)

def senha_redefinida_redirect(request):
    messages.success(request, "Senha redefinida com sucesso! Faça login para continuar.")
    return redirect('login')

@login_required
@staff_member_required
def gerenciar_tags(request):
    if request.method == 'POST':
        action = request.POST.get('action')
        tag_id = request.POST.get('tag_id')

        if action == 'add':
            form = TagForm(request.POST)
            if form.is_valid():
                form.save()
                messages.success(request, f"Tag '{form.cleaned_data['nome']}' criada com sucesso.")
            else:
                messages.error(request, "Erro ao criar a tag: " + form.errors.as_text())

        elif action == 'edit' and tag_id:
            tag = get_object_or_404(Tag, id=tag_id)
            form = TagForm(request.POST, instance=tag)
            if form.is_valid():
                form.save()
                messages.success(request, f"Tag '{form.cleaned_data['nome']}' renomeada com sucesso.")
            else:
                messages.error(request, "Erro ao renomear: " + form.errors.as_text())

        elif action == 'delete' and tag_id:
            try:
                tag = get_object_or_404(Tag, id=tag_id)
                tag_nome = tag.nome
                tag.delete()
                messages.success(request, f"Tag '{tag_nome}' excluída com sucesso!")
            except Exception as e:
                messages.error(request, f"Erro ao excluir a tag: {e}")
        
        elif action == 'update_children' and tag_id:
            parent_tag = get_object_or_404(Tag, id=tag_id)
            child_ids = request.POST.getlist('children')

            child_tags = Tag.objects.filter(id__in=child_ids)
            parent_tag.children.set(child_tags)
            parent_tag.children.exclude(id__in=child_ids).update(parent=None)

            messages.success(request, f"Subcategorias de '{parent_tag.nome}' atualizadas com sucesso.")

        return redirect('gerenciar_tags')

    all_tags = Tag.objects.all().order_by('nome')
    parent_tags = all_tags.filter(parent__isnull=True).prefetch_related('children')
    
    form = TagForm()
    
    context = {
        'parent_tags': parent_tags,
        'all_tags': all_tags,
        'form': form,
    }
    return render(request, 'core/gerenciar_tags.html', context)

@login_required
@require_POST
def adicionar_avaliacao(request, slug):
    empresa = get_object_or_404(Empresa, slug=slug)
    form = AvaliacaoForm(request.POST)

    if form.is_valid():
        try:
            avaliacao = form.save(commit=False)
            avaliacao.empresa = empresa
            avaliacao.user = request.user
            avaliacao.save()
            messages.success(request, "Obrigado pela sua avaliação!")
        except IntegrityError:
            messages.error(request, "Você já avaliou esta empresa.")
    else:
        messages.error(request, "Houve um erro no seu formulário. Por favor, tente novamente.")

    return redirect('empresa_detalhe', slug=empresa.slug)

@login_required
@require_POST
def deletar_avaliacao(request, avaliacao_id):
    avaliacao = get_object_or_404(Avaliacao, id=avaliacao_id)

    if request.user != avaliacao.user and not request.user.is_superuser:
        return JsonResponse({'status': 'error', 'message': 'Permissão negada.'}, status=403)

    avaliacao.delete()

    messages.success(request, "Sua avaliação foi removida com sucesso.")
    return JsonResponse({'status': 'success'})

@login_required
@require_POST
def toggle_favorito(request, slug):
    empresa = get_object_or_404(Empresa, slug=slug)
    perfil = request.user.perfil

    if empresa in perfil.favoritos.all():
        perfil.favoritos.remove(empresa)
        is_favorito = False
    else:
        perfil.favoritos.add(empresa)
        is_favorito = True

    return JsonResponse({'status': 'ok', 'is_favorito': is_favorito})


@login_required
def listar_favoritos(request):
    perfil = request.user.perfil
    
    empresas_favoritas = get_base_empresas_queryset().filter(
        id__in=perfil.favoritos.values_list('id', flat=True)
    ).order_by('-data_cadastro')

    paginator = Paginator(empresas_favoritas, 12)
    page_obj = paginator.get_page(request.GET.get('page') or 1)

    return render(request, 'core/listar_favoritos.html', {
        'page_obj': page_obj
    })

@login_required
def gerador_qrcode_view(request):
    all_tags = list(Tag.objects.order_by('nome').values('id', 'nome'))
    all_cidades = list(
        Empresa.objects.exclude(cidade__isnull=True).exclude(cidade__exact='')
        .order_by('cidade').values_list('cidade', flat=True).distinct()
    )
    
    context = {
        'all_tags': all_tags,
        'all_cidades': all_cidades,
    }
    return render(request, 'core/gerador_qrcode.html', context)

@login_required
@require_POST
def salvar_tema_preferido(request):
    try:
        data = json.loads(request.body)
        tema_escolhido = data.get('theme')

        allowed_themes = [choice[0] for choice in PerfilUsuario.TEMA_ESCOLHAS]
        if tema_escolhido not in allowed_themes:
            return JsonResponse({'status': 'error', 'message': 'Tema inválido'}, status=400)

        perfil = request.user.perfil
        perfil.tema_preferido = tema_escolhido
        perfil.save(update_fields=['tema_preferido'])
        
        return JsonResponse({'status': 'ok'})
    except (json.JSONDecodeError, AttributeError):
        return JsonResponse({'status': 'error', 'message': 'Requisição inválida'}, status=400)